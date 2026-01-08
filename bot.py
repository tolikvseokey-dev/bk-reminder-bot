import os
import json
import uuid
import time
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional, Tuple

import pytz
import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from apscheduler.schedulers.background import BackgroundScheduler

try:
    from openpyxl import load_workbook
except Exception as e:
    raise RuntimeError("Не установлен openpyxl. Добавь в requirements.txt строку: openpyxl") from e


# ================== ВЕРСИЯ ==================
BOT_VERSION = "topic-locked-storage-no-exit-no-reload-stop-admin-2026-01-08-07"


# ================== НАСТРОЙКИ ==================
BOT_TOKEN = os.environ.get("BOT_TOKEN", "").strip()
DATA_FILE = "reminders.json"

TZ_NAME = os.environ.get("BOT_TZ", "Europe/Moscow")
TZ = pytz.timezone(TZ_NAME)

DATE_PICK_DAYS = int(os.environ.get("DATE_PICK_DAYS", "21"))

AUTO_DELETE_AFTER_HOURS = int(os.environ.get("AUTO_DELETE_AFTER_HOURS", "24"))
CLEANUP_INTERVAL_MINUTES = int(os.environ.get("CLEANUP_INTERVAL_MINUTES", "1"))

STORAGE_FILE_ENV = os.environ.get("STORAGE_FILE", "").strip()

ADMIN_USERNAME = "AnatoliiOsin"   # только он видит админ-кнопки

if not BOT_TOKEN:
    raise RuntimeError("Не задан BOT_TOKEN. Добавь переменную окружения BOT_TOKEN в панели хостинга (Bothost).")

bot = telebot.TeleBot(BOT_TOKEN, parse_mode="HTML")
scheduler = BackgroundScheduler(timezone=TZ)
scheduler.start()

states: Dict[int, Dict[str, Any]] = {}


# ================== HELPERS (ADMIN / TOPICS) ==================
def is_admin_user(user) -> bool:
    try:
        return (user.username or "").strip() == ADMIN_USERNAME
    except Exception:
        return False


def chat_is_group(chat) -> bool:
    try:
        return chat.type in ("group", "supergroup")
    except Exception:
        return False


def get_thread_id_from_message(message) -> Optional[int]:
    try:
        return getattr(message, "message_thread_id", None)
    except Exception:
        return None


def get_thread_id_from_call(call) -> Optional[int]:
    try:
        return getattr(call.message, "message_thread_id", None)
    except Exception:
        return None


# ================== ХРАНЕНИЕ (JSON) ==================
def load_data() -> Dict[str, Any]:
    """
    Структура:
    {
      "reminders": [ ... ],
      "chat_settings": {
          "<chat_id>": {
              "allowed_thread_id": 123
          }
      }
    }
    """
    if not os.path.exists(DATA_FILE):
        return {"reminders": [], "chat_settings": {}}

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            return {"reminders": [], "chat_settings": {}}

    if "reminders" not in data or not isinstance(data["reminders"], list):
        data["reminders"] = []
    if "chat_settings" not in data or not isinstance(data["chat_settings"], dict):
        data["chat_settings"] = {}
    return data


def save_data(data: Dict[str, Any]) -> None:
    if "reminders" not in data:
        data["reminders"] = []
    if "chat_settings" not in data:
        data["chat_settings"] = {}
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def now_tz() -> datetime:
    return datetime.now(TZ)


def dt_from_iso(iso_str: str) -> Optional[datetime]:
    try:
        dt = datetime.fromisoformat(iso_str)
        if dt.tzinfo is None:
            dt = TZ.localize(dt)
        else:
            dt = dt.astimezone(TZ)
        return dt
    except Exception:
        return None


def dt_to_iso(dt: datetime) -> str:
    if dt.tzinfo is None:
        dt = TZ.localize(dt)
    else:
        dt = dt.astimezone(TZ)
    return dt.isoformat()


def add_reminder_to_store(rem: Dict[str, Any]) -> None:
    data = load_data()
    data["reminders"].append(rem)
    save_data(data)


def get_chat_reminders(chat_id: int) -> List[Dict[str, Any]]:
    data = load_data()
    items = [r for r in data.get("reminders", []) if int(r.get("chat_id", 0)) == int(chat_id)]

    changed = False
    for r in items:
        dt = dt_from_iso(r.get("event_dt", ""))
        if dt:
            new_iso = dt_to_iso(dt)
            if r.get("event_dt") != new_iso:
                r["event_dt"] = new_iso
                changed = True

    items.sort(key=lambda r: r.get("event_dt", ""))

    if changed:
        all_data = load_data()
        by_id = {r.get("id"): r for r in items if r.get("id")}
        for i, r in enumerate(all_data.get("reminders", [])):
            rid = r.get("id")
            if rid in by_id:
                all_data["reminders"][i] = by_id[rid]
        save_data(all_data)

    return items


def get_allowed_thread_id(chat_id: int) -> Optional[int]:
    data = load_data()
    st = (data.get("chat_settings") or {}).get(str(chat_id), {})
    tid = st.get("allowed_thread_id")
    try:
        return int(tid) if tid is not None else None
    except Exception:
        return None


def set_allowed_thread_id(chat_id: int, thread_id: int) -> None:
    data = load_data()
    cs = data.setdefault("chat_settings", {})
    cs.setdefault(str(chat_id), {})["allowed_thread_id"] = int(thread_id)
    save_data(data)


def clear_allowed_thread_id(chat_id: int) -> None:
    data = load_data()
    cs = data.setdefault("chat_settings", {})
    if str(chat_id) in cs:
        cs[str(chat_id)].pop("allowed_thread_id", None)
    save_data(data)


def in_allowed_topic_for_message(message) -> bool:
    """
    Правило:
    - если не группа/супергруппа -> True
    - если allowed_thread_id не задан -> разрешаем только админу (чтобы он мог закрепить тему)
    - если задан -> сообщение должно быть в этой теме (message_thread_id == allowed)
    """
    if not chat_is_group(message.chat):
        return True

    allowed = get_allowed_thread_id(message.chat.id)
    if allowed is None:
        return is_admin_user(message.from_user)

    tid = get_thread_id_from_message(message)
    return tid == allowed


def in_allowed_topic_for_call(call) -> bool:
    if not chat_is_group(call.message.chat):
        return True

    allowed = get_allowed_thread_id(call.message.chat.id)
    if allowed is None:
        return is_admin_user(call.from_user)

    tid = get_thread_id_from_call(call)
    return tid == allowed


def send_locked(chat_id: int, text: str, reply_markup=None, disable_web_page_preview: bool = False, fallback_thread_id: Optional[int] = None):
    """
    Отправка сообщений:
    - Если чат групповой и тема закреплена -> ВСЕГДА отправляем в закреплённую тему.
    - Если тема не закреплена -> если передан fallback_thread_id, отправим туда (удобно для шага закрепления),
      иначе обычным способом.
    """
    allowed = get_allowed_thread_id(chat_id)
    try:
        if allowed is not None:
            return bot.send_message(
                chat_id,
                text,
                reply_markup=reply_markup,
                disable_web_page_preview=disable_web_page_preview,
                message_thread_id=allowed
            )
        if fallback_thread_id is not None:
            return bot.send_message(
                chat_id,
                text,
                reply_markup=reply_markup,
                disable_web_page_preview=disable_web_page_preview,
                message_thread_id=fallback_thread_id
            )
        return bot.send_message(
            chat_id,
            text,
            reply_markup=reply_markup,
            disable_web_page_preview=disable_web_page_preview
        )
    except Exception:
        # если отправка в тему упала — попробуем без темы
        return bot.send_message(
            chat_id,
            text,
            reply_markup=reply_markup,
            disable_web_page_preview=disable_web_page_preview
        )


# ================== INLINE МЕНЮ ==================
def kb_main_inline(user=None) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("📌 Напоминания", callback_data="nav_reminders"))
    kb.row(InlineKeyboardButton("📚 Полезная информация", callback_data="nav_useful"))
    kb.row(InlineKeyboardButton("🧊 Сроки хранения (поиск)", callback_data="nav_storage"))
    kb.row(InlineKeyboardButton("ℹ️ О боте", callback_data="nav_about"))

    # Админ-блок (только AnatoliiOsin)
    if user is not None and is_admin_user(user):
        kb.row(InlineKeyboardButton("📌 Закрепить эту тему", callback_data="admin_pin_topic"))
        kb.row(InlineKeyboardButton("🛑 Остановить бота", callback_data="admin_stop_bot"))
    return kb


def kb_reminders_inline(user=None) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("➕ Добавить напоминание", callback_data="rem_add"))
    kb.row(InlineKeyboardButton("📋 Все напоминания", callback_data="rem_list"))
    kb.row(InlineKeyboardButton("⬅️ Назад", callback_data="nav_main"))
    return kb


def kb_cancel_inline() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    kb.row(InlineKeyboardButton("⬅️ Назад в меню", callback_data="nav_main"))
    return kb


# ================== ПОЛЕЗНАЯ ИНФОРМАЦИЯ (INLINE) ==================
USEFUL_LINKS = {
    "rm_schedule": "https://docs.google.com/spreadsheets/d/1ZXCllmYkqmP6y9HRnYm0_2D2f63haeU-vI2gylnL6Pg/edit?usp=drive_link",
    "vacations": "https://docs.google.com/spreadsheets/d/12SEymi_QNwSJ8agRBzXc1UZCfNhabtiLX07KxEsmpzQ/edit?usp=drive_link",
    "ato": "https://docs.google.com/spreadsheets/d/1IiKxS9Tf6oHUJJDhfozvWdbhC9wOZPzapflYv612Du0/edit",
    "dynamics": "https://docs.google.com/spreadsheets/d/1HhgNo3mfd8LrdfBPU2sjVatA-fboBf75387Ryd-qVUg/edit?gid=2086138160#gid=2086138160",
    "roster": "https://docs.google.com/spreadsheets/d/1vwPI_SPnjX5wPI6tu4jAFXSWFubjBQEO56kuCMysL_4/edit?usp=drive_link",
    "contacts": "https://docs.google.com/spreadsheets/d/1P5GbNMQD0A3OWh6GxLAYJDlgC92H95uo/edit?gid=2031453167#gid=2031453167",
    "protocol_rm": "https://docs.google.com/spreadsheets/d/1dBZzfanIbtjgp2sFDzU441Wv6ghT-bryQ19wc034Ye4/edit",
    "protocol_directors": "https://docs.google.com/spreadsheets/d/1cEMp3_84LuXrffAgqAOQq9kG8k-Ks8ev5k3Xo3QR-qo/edit",
}

GROUPS_TEXT = """Группы

Витрины
https://t.me/+9hdkceSRFdU4MmZi

Кофе-бар
https://t.me/+rAM0-VID0Gg0NmUy

Персонал
https://t.me/+ZcNnavnmJQlkZDAy

Цели
https://t.me/+SkzL_Xit6ypkMmZi

Айти вопросы
https://t.me/+oMzRrI1DzGlkNDVi

Логистика
https://t.me/+CsD1pmYTTnQ5NDdi

Технические вопросы
https://t.me/+0dm4nBMj3LVlMGYy

Качество ЛБК
https://t.me/+9WmWOSrjBxs1N2Uy

Заказы РЦ и ФК
Нет ссылки

Выход на стажировку
https://t.me/+fa-ESZUYflA0ZThi

Обратная связь ЕДА
https://t.me/+3mipmXTpud5kZWVi

5 pillars
https://t.me/+f_YYEYz1rfc4NjAy

Курьеры кадры
https://t.me/+E7w0LSi4ltBlZjJi

Поиск продукции
https://t.me/+Yw-opolA0tc5ZTY6

Корп академия
https://t.me/+uhlNZjfkeZE0NGYy

Обучение БК
https://t.me/+GU5oGnyjdgc5OTMy

Важная информация
https://t.me/+QB6nQlAno9xhZTQy

Пожарная безопасность
https://t.me/+l2rMTNe2I_VkMjNi
"""


def kb_useful_inline() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("🗓 Расписание РМ", url=USEFUL_LINKS["rm_schedule"]))
    kb.row(InlineKeyboardButton("🌴 График отпусков", url=USEFUL_LINKS["vacations"]))
    kb.row(InlineKeyboardButton("📊 АТО", url=USEFUL_LINKS["ato"]))
    kb.row(InlineKeyboardButton("🔗 Ссылки на группы", callback_data="ui_groups"))
    kb.row(InlineKeyboardButton("📈 Динамика", url=USEFUL_LINKS["dynamics"]))
    kb.row(InlineKeyboardButton("🧾 Ростер", url=USEFUL_LINKS["roster"]))
    kb.row(InlineKeyboardButton("☎️ Контакт-лист", url=USEFUL_LINKS["contacts"]))
    kb.row(InlineKeyboardButton("📝 Протокол собрания", callback_data="ui_protocol"))
    kb.row(InlineKeyboardButton("⬅️ Назад", callback_data="nav_main"))
    return kb


def kb_protocol_inline() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("👔 РМ", url=USEFUL_LINKS["protocol_rm"]))
    kb.row(InlineKeyboardButton("🧑‍💼 Директора", url=USEFUL_LINKS["protocol_directors"]))
    kb.row(InlineKeyboardButton("⬅️ Назад", callback_data="nav_useful"))
    return kb


# ================== INLINE ПИКЕРЫ ДАТЫ/ВРЕМЕНИ ==================
def build_date_picker() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    today = now_tz().date()

    buttons = []
    for i in range(DATE_PICK_DAYS):
        d = today + timedelta(days=i)
        dow = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"][d.weekday()]
        text = d.strftime("%d.%m") + f" ({dow})"
        buttons.append(InlineKeyboardButton(text, callback_data=f"date|{d.isoformat()}"))

    for i in range(0, len(buttons), 2):
        kb.row(*buttons[i:i + 2])

    kb.row(InlineKeyboardButton("✍️ Ввести дату вручную", callback_data="date_manual"))
    kb.row(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return kb


def build_time_picker() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    common = ["09:00", "12:00", "15:00", "18:00", "21:00"]
    for i in range(0, len(common), 2):
        row = []
        for t in common[i:i + 2]:
            row.append(InlineKeyboardButton(t, callback_data=f"time|{t}"))
        kb.row(*row)

    kb.row(InlineKeyboardButton("✍️ Ввести время вручную", callback_data="time_manual"))
    kb.row(InlineKeyboardButton("❌ Отмена", callback_data="cancel"))
    return kb


def validate_time_hhmm(s: str) -> bool:
    try:
        datetime.strptime(s, "%H:%M")
        return True
    except ValueError:
        return False


def format_event_dt(iso_str: str) -> str:
    dt = dt_from_iso(iso_str)
    if not dt:
        return "неизвестная дата"
    return dt.strftime("%d.%m.%Y %H:%M")


# ================== ПЛАНИРОВЩИК НАПОМИНАНИЙ ==================
def schedule_reminder_jobs(reminder: Dict[str, Any]) -> None:
    rem_id = reminder["id"]
    chat_id = int(reminder["chat_id"])
    title = reminder["title"]

    event_dt = dt_from_iso(reminder["event_dt"])
    if not event_dt:
        return

    thread_id = reminder.get("thread_id")
    try:
        thread_id = int(thread_id) if thread_id is not None else None
    except Exception:
        thread_id = None

    for kind, delta, label in [
        ("24h", timedelta(hours=24), "за 24 часа"),
        ("1h", timedelta(hours=1), "за 1 час"),
    ]:
        run_at = event_dt - delta
        job_id = f"{rem_id}_{kind}"

        if run_at <= now_tz():
            try:
                scheduler.remove_job(job_id)
            except Exception:
                pass
            continue

        def _send(chat_id=chat_id, title=title, event_dt=event_dt, label=label, thread_id=thread_id):
            send_locked(
                chat_id,
                f"⏰ Напоминание ({label})\n"
                f"<b>{title}</b>\n"
                f"📅 Событие: <b>{event_dt.strftime('%d.%m.%Y %H:%M')}</b>",
                fallback_thread_id=thread_id
            )

        scheduler.add_job(
            _send,
            trigger="date",
            run_date=run_at,
            id=job_id,
            replace_existing=True,
            misfire_grace_time=60 * 10
        )


def reschedule_all_from_store() -> None:
    data = load_data()
    for r in data.get("reminders", []):
        dt = dt_from_iso(r.get("event_dt", ""))
        if dt:
            r["event_dt"] = dt_to_iso(dt)
        schedule_reminder_jobs(r)
    save_data(data)


def cleanup_expired() -> None:
    data = load_data()
    reminders = data.get("reminders", [])
    if not reminders:
        return

    cutoff = now_tz() - timedelta(hours=AUTO_DELETE_AFTER_HOURS)
    keep: List[Dict[str, Any]] = []
    removed_ids: List[str] = []

    for r in reminders:
        dt = dt_from_iso(r.get("event_dt", ""))
        if not dt:
            removed_ids.append(r.get("id", ""))
            continue

        if dt < cutoff:
            removed_ids.append(r.get("id", ""))
        else:
            r["event_dt"] = dt_to_iso(dt)
            keep.append(r)

    if removed_ids:
        for rid in removed_ids:
            if not rid:
                continue
            for kind in ("24h", "1h"):
                try:
                    scheduler.remove_job(f"{rid}_{kind}")
                except Exception:
                    pass

        data["reminders"] = keep
        save_data(data)


reschedule_all_from_store()
scheduler.add_job(
    cleanup_expired,
    trigger="interval",
    minutes=CLEANUP_INTERVAL_MINUTES,
    id="cleanup_expired",
    replace_existing=True
)


# ================== БАЗА СРОКОВ ХРАНЕНИЯ (XLSX) ==================
def _script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def find_storage_file() -> Optional[str]:
    if STORAGE_FILE_ENV:
        p = STORAGE_FILE_ENV
        if not os.path.isabs(p):
            p = os.path.join(_script_dir(), p)
        if os.path.exists(p):
            return p

    candidates = [
        "storage.xlsx",
        "Storage.xlsx",
        "Storage .xlsx",
        "Storage  .xlsx",
        "Меню БК без картинок .xlsx",
        "Меню БК без картинок.xlsx",
    ]
    for name in candidates:
        p = os.path.join(_script_dir(), name)
        if os.path.exists(p):
            return p
    return None


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() == "none":
        return ""
    return s


def _norm(s: str) -> str:
    return " ".join((s or "").strip().lower().split())


StorageRow = Dict[str, Any]

STORAGE_DB: List[StorageRow] = []
STORAGE_READY: bool = False
STORAGE_SOURCE_PATH: str = ""

# Канон (человеческие подписи для вывода)
H_NAME = "Наименование"
H_OUT = "Выход"
H_SHELF = "Срок хранения"
H_TEMP = "Рекомендуемая температура отдачи блюд"
H_MARK = "Маркировка на витрине"
H_LAYOUT = "Стандарт выкладки"
H_PACK = "Упаковка с собой/доставка"

BASE_ALWAYS = [H_OUT, H_SHELF, H_TEMP]          # всегда выводим, пустое -> —
OPTIONAL_IF_FILLED = [H_MARK, H_LAYOUT]         # выводим только если заполнено


def _canonical_header(raw: str) -> Optional[str]:
    """
    Алиасы под твой файл.
    Важно: H_MARK и H_LAYOUT — разные поля.
    """
    t = _norm(raw)
    if not t:
        return None

    if "наимен" in t or t == "название" or "наименов" in t:
        return H_NAME

    if "выход" in t:
        return H_OUT

    if "срок" in t or "реализац" in t:
        return H_SHELF

    if "температур" in t and ("отдач" in t or "блюд" in t):
        return H_TEMP

    if "маркиров" in t:
        return H_MARK

    if "стандарт" in t and "выклад" in t:
        return H_LAYOUT

    if "упаков" in t or "с собой" in t or "достав" in t:
        return H_PACK

    return None


def _guess_header_row(ws, max_rows: int = 10, max_cols: int = 30) -> int:
    best_row = 1
    best_score = -1

    for r in range(1, max_rows + 1):
        seen = set()
        score = 0
        for c in range(1, max_cols + 1):
            h_raw = _cell_str(ws.cell(r, c).value)
            canon = _canonical_header(h_raw)
            if not canon or canon in seen:
                continue
            seen.add(canon)

        if H_NAME in seen:
            score += 5
        if H_OUT in seen:
            score += 2
        if H_SHELF in seen:
            score += 2
        if H_TEMP in seen:
            score += 1
        if H_MARK in seen:
            score += 1
        if H_LAYOUT in seen:
            score += 1
        if H_PACK in seen:
            score += 1

        if score > best_score:
            best_score = score
            best_row = r

    return best_row if best_score >= 5 else 1


def load_storage_db() -> Tuple[int, List[str]]:
    global STORAGE_DB, STORAGE_READY, STORAGE_SOURCE_PATH

    path = find_storage_file()
    STORAGE_DB = []
    STORAGE_READY = False
    STORAGE_SOURCE_PATH = path or ""

    if not path:
        return 0, []

    wb = load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        header_row = _guess_header_row(ws)

        col_by_header: Dict[str, int] = {}
        for col in range(1, 31):
            h_raw = _cell_str(ws.cell(row=header_row, column=col).value)
            canon = _canonical_header(h_raw)
            if canon and canon not in col_by_header:
                col_by_header[canon] = col

        name_col = col_by_header.get(H_NAME, 1)
        sheet_has_pack = H_PACK in col_by_header

        cols = {
            H_OUT: col_by_header.get(H_OUT),
            H_SHELF: col_by_header.get(H_SHELF),
            H_TEMP: col_by_header.get(H_TEMP),
            H_MARK: col_by_header.get(H_MARK),
            H_LAYOUT: col_by_header.get(H_LAYOUT),
            H_PACK: col_by_header.get(H_PACK),  # может быть None
        }

        for row in range(header_row + 1, ws.max_row + 1):
            name = _cell_str(ws.cell(row=row, column=name_col).value)
            if not name:
                continue

            fields: Dict[str, str] = {}
            any_field = False

            for h, c in cols.items():
                v = _cell_str(ws.cell(row=row, column=c).value) if c else ""
                fields[h] = v
                if v:
                    any_field = True

            if not any_field:
                continue

            STORAGE_DB.append({
                "category": sheet_name,
                "name": name,
                "name_lc": name.lower(),
                "fields": fields,
                "sheet_has_pack": sheet_has_pack,
            })

    STORAGE_READY = len(STORAGE_DB) > 0
    return len(STORAGE_DB), sheet_names


def storage_search(query: str, limit: int = 12) -> List[StorageRow]:
    q = (query or "").strip().lower()
    if not q:
        return []

    hits = [row for row in STORAGE_DB if q in row["name_lc"]]

    if not hits:
        parts = [p for p in q.split() if p]
        if parts:
            hits = [row for row in STORAGE_DB if all(p in row["name_lc"] for p in parts)]

    return hits[:limit]


def format_storage_row(row: StorageRow) -> str:
    category = row.get("category", "")
    name = row.get("name", "")
    fields: Dict[str, str] = row.get("fields", {}) or {}
    sheet_has_pack: bool = bool(row.get("sheet_has_pack", False))

    lines = []
    if category:
        lines.append(f"📂 <b>{category}</b>")
    if name:
        lines.append(f"\n<b>{name}</b>")

    for h in BASE_ALWAYS:
        v = _cell_str(fields.get(h, ""))
        lines.append(f"\n<b>{h}:</b>\n{v if v else '—'}")

    if sheet_has_pack:
        v = _cell_str(fields.get(H_PACK, ""))
        lines.append(f"\n<b>{H_PACK}:</b>\n{v if v else '—'}")

    for h in OPTIONAL_IF_FILLED:
        v = _cell_str(fields.get(h, ""))
        if v:
            lines.append(f"\n<b>{h}:</b>\n{v}")

    return "\n".join(lines).strip()


# ====== КЛАВЫ ДЛЯ СРОКОВ ХРАНЕНИЯ (без Exit и без Reload) ======
def kb_storage_start() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("🔎 Новый поиск", callback_data="storage_newsearch"))
    kb.row(InlineKeyboardButton("⬅️ В меню", callback_data="nav_main"))
    return kb


def kb_storage_pick_list(results: List[StorageRow]) -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    for i, row in enumerate(results[:8]):
        title = row.get("name", "")
        if len(title) > 40:
            title = title[:40] + "…"
        kb.row(InlineKeyboardButton(f"{i + 1}) {title}", callback_data=f"storage_pick|{i}"))
    kb.row(InlineKeyboardButton("🔎 Новый поиск", callback_data="storage_newsearch"))
    kb.row(InlineKeyboardButton("⬅️ В меню", callback_data="nav_main"))
    return kb


def kb_storage_after_result() -> InlineKeyboardMarkup:
    kb = InlineKeyboardMarkup()
    kb.row(InlineKeyboardButton("🔎 Новый поиск", callback_data="storage_newsearch"))
    kb.row(InlineKeyboardButton("⬅️ В меню", callback_data="nav_main"))
    return kb


# ================== STATE HELPERS ==================
def clear_user_state(user_id: int) -> None:
    states.pop(user_id, None)


def clear_storage_mode(user_id: int) -> None:
    st = states.get(user_id)
    if not st:
        return
    if st.get("mode") == "storage_search":
        clear_user_state(user_id)


# ================== УТИЛИТА: УБРАТЬ СТАРУЮ REPLY-КЛАВУ ==================
def remove_old_keyboard(chat_id: int, thread_id: Optional[int] = None) -> None:
    send_locked(chat_id, "Обновил меню ✅", reply_markup=ReplyKeyboardRemove(), fallback_thread_id=thread_id)


# ================== /start /menu ==================
@bot.message_handler(commands=["start", "menu"])
def start_cmd(message):
    # topic-lock filter
    if not in_allowed_topic_for_message(message):
        return

    clear_user_state(message.from_user.id)
    remove_old_keyboard(message.chat.id, get_thread_id_from_message(message))

    allowed = get_allowed_thread_id(message.chat.id) if chat_is_group(message.chat) else None
    if chat_is_group(message.chat) and allowed is None and not is_admin_user(message.from_user):
        # не спамим в группах: до закрепления темы отвечаем только админу
        return

    intro = "Главное меню 👇\n" f"<i>Версия: {BOT_VERSION}</i>"
    if chat_is_group(message.chat) and allowed is None and is_admin_user(message.from_user):
        intro += (
            "\n\n⚠️ <b>Тема ещё не закреплена.</b>\n"
            "Перейди в нужную тему и нажми «📌 Закрепить эту тему» — после этого бот будет жить только там."
        )

    send_locked(message.chat.id, intro, reply_markup=kb_main_inline(message.from_user), fallback_thread_id=get_thread_id_from_message(message))


# ================== КОМАНДЫ ТОЛЬКО ДЛЯ АДМИНА ==================
@bot.message_handler(commands=["storage_reload"])
def admin_storage_reload(message):
    if not is_admin_user(message.from_user):
        return
    # разрешаем админу даже если тема не закреплена, но работаем в текущей теме
    if not in_allowed_topic_for_message(message):
        return
    count, sheets = load_storage_db()
    tid = get_thread_id_from_message(message)
    if count == 0:
        send_locked(
            message.chat.id,
            "❌ Не нашёл файл базы или база пустая.\n"
            "Проверь, что xlsx лежит рядом с bot.py или задай STORAGE_FILE.",
            reply_markup=kb_main_inline(message.from_user),
            fallback_thread_id=tid
        )
        return
    send_locked(
        message.chat.id,
        f"✅ База перезагружена: <b>{count}</b> строк.\n"
        f"Листы: {', '.join(sheets)}",
        reply_markup=kb_main_inline(message.from_user),
        fallback_thread_id=tid
    )


@bot.message_handler(commands=["topic_clear"])
def admin_topic_clear(message):
    if not is_admin_user(message.from_user):
        return
    if not in_allowed_topic_for_message(message):
        return
    clear_allowed_thread_id(message.chat.id)
    send_locked(
        message.chat.id,
        "✅ Привязка к теме сброшена. Теперь снова нужно закрепить тему кнопкой «📌 Закрепить эту тему».",
        reply_markup=kb_main_inline(message.from_user),
        fallback_thread_id=get_thread_id_from_message(message)
    )


# ================== ПОДХВАТ СТАРЫХ КНОПОК (если их нажмут) ==================
@bot.message_handler(func=lambda m: (m.text or "").strip() in {
    "📌 Напоминания", "📚 Полезная информация", "ℹ️ О боте",
    "➕ Добавить напоминание", "📋 Все напоминания", "⬅️ Назад"
})
def legacy_buttons_handler(message):
    if not in_allowed_topic_for_message(message):
        return
    clear_user_state(message.from_user.id)
    remove_old_keyboard(message.chat.id, get_thread_id_from_message(message))
    send_locked(message.chat.id, "Перешли на новое меню (inline) 👇", reply_markup=kb_main_inline(message.from_user),
                fallback_thread_id=get_thread_id_from_message(message))


# ================== NAV + ADMIN CALLBACKS ==================
@bot.callback_query_handler(func=lambda call: call.data.startswith(("nav_", "admin_")))
def nav_callbacks(call):
    chat_id = call.message.chat.id
    user_id = call.from_user.id
    data = call.data

    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass

    # topic-lock filter
    if not in_allowed_topic_for_call(call):
        return

    if data == "admin_pin_topic":
        if not is_admin_user(call.from_user):
            return
        if not chat_is_group(call.message.chat):
            send_locked(chat_id, "Эта кнопка нужна только в группах с темами.", reply_markup=kb_main_inline(call.from_user))
            return

        tid = get_thread_id_from_call(call)
        if tid is None:
            send_locked(
                chat_id,
                "⚠️ Я не вижу ID темы.\n"
                "Открой <b>нужную тему</b> (Forum Topic) и нажми «📌 Закрепить эту тему» там.",
                reply_markup=kb_main_inline(call.from_user)
            )
            return

        set_allowed_thread_id(chat_id, tid)
        clear_user_state(user_id)
        send_locked(
            chat_id,
            f"✅ Готово! Закрепил эту тему.\n\n"
            f"Теперь я буду отвечать <b>только здесь</b> и игнорировать другие темы.\n"
            f"<i>thread_id={tid}</i>",
            reply_markup=kb_main_inline(call.from_user),
            fallback_thread_id=tid
        )
        return

    if data == "admin_stop_bot":
        if not is_admin_user(call.from_user):
            return
        clear_user_state(user_id)
        send_locked(chat_id, "🛑 Останавливаю бота…", reply_markup=None, fallback_thread_id=get_thread_id_from_call(call))
        try:
            scheduler.shutdown(wait=False)
        except Exception:
            pass
        # Жестко завершаем процесс — на хостинге он обычно перезапустится супервизором, если настроено.
        os._exit(0)

    if data == "nav_main":
        clear_user_state(user_id)
        try:
            bot.edit_message_text("Главное меню 👇", chat_id, call.message.message_id, reply_markup=kb_main_inline(call.from_user))
        except Exception:
            send_locked(chat_id, "Главное меню 👇", reply_markup=kb_main_inline(call.from_user), fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "nav_reminders":
        clear_user_state(user_id)
        try:
            bot.edit_message_text("📌 <b>Напоминания</b> — выбери действие:", chat_id, call.message.message_id,
                                  reply_markup=kb_reminders_inline(call.from_user))
        except Exception:
            send_locked(chat_id, "📌 <b>Напоминания</b> — выбери действие:", reply_markup=kb_reminders_inline(call.from_user),
                        fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "nav_useful":
        clear_user_state(user_id)
        try:
            bot.edit_message_text("📚 <b>Полезная информация</b> — выбери пункт:", chat_id, call.message.message_id,
                                  reply_markup=kb_useful_inline())
        except Exception:
            send_locked(chat_id, "📚 <b>Полезная информация</b> — выбери пункт:", reply_markup=kb_useful_inline(),
                        fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "nav_storage":
        # если тема закреплена — вход разрешен только из неё (фильтр выше уже отработал)
        # если не закреплена — разрешаем вход только админу, и только из темы (tid != None)
        if chat_is_group(call.message.chat) and get_allowed_thread_id(chat_id) is None and not is_admin_user(call.from_user):
            return

        tid = get_thread_id_from_call(call)

        if chat_is_group(call.message.chat) and get_allowed_thread_id(chat_id) is None:
            # админ ещё не закрепил тему
            if tid is None:
                send_locked(chat_id, "Открой тему и нажми «📌 Закрепить эту тему» — потом заходи в поиск.",
                            reply_markup=kb_main_inline(call.from_user))
                return

        if not STORAGE_READY:
            send_locked(
                chat_id,
                "🧊 <b>Сроки хранения</b>\n\n"
                "База не загружена или пустая.\n"
                "Проверь файл рядом с bot.py или попроси админа выполнить /storage_reload",
                reply_markup=kb_storage_start(),
                fallback_thread_id=tid
            )
            return

        states[user_id] = {"mode": "storage_search", "chat_id": chat_id, "thread_id": tid}
        send_locked(
            chat_id,
            "🧊 <b>Сроки хранения — поиск</b>\n\n"
            "Введи название продукта (можно часть слова).\n"
            "Пример: <i>омлет</i>, <i>песто</i>, <i>суп</i>",
            reply_markup=kb_storage_start(),
            fallback_thread_id=tid
        )
        return

    if data == "nav_about":
        clear_user_state(user_id)
        allowed = get_allowed_thread_id(chat_id)
        text = (
            "ℹ️ <b>О боте</b>\n\n"
            "• Напоминания: добавление и список\n"
            "• Полезная информация: ссылки/материалы\n"
            "• Сроки хранения: поиск по Excel базе\n"
            "• Режим темы: бот живёт только в одной теме (после закрепления)\n\n"
            f"🕒 Таймзона: <b>{TZ_NAME}</b>\n"
            f"🧹 Автоудаление напоминаний: <b>{AUTO_DELETE_AFTER_HOURS} ч</b> после события\n"
            f"🧊 База сроков хранения: <b>{'загружена' if STORAGE_READY else 'не загружена'}</b>\n"
            f"📌 Закреплённая тема: <b>{allowed if allowed is not None else 'не задана'}</b>\n"
            f"🔖 Версия: <b>{BOT_VERSION}</b>"
        )
        try:
            bot.edit_message_text(text, chat_id, call.message.message_id, reply_markup=kb_main_inline(call.from_user))
        except Exception:
            send_locked(chat_id, text, reply_markup=kb_main_inline(call.from_user), fallback_thread_id=get_thread_id_from_call(call))
        return


# ================== CALLBACKS (полезная информация) ==================
@bot.callback_query_handler(func=lambda call: call.data.startswith("ui_"))
def callbacks_useful(call):
    # topic-lock filter
    if not in_allowed_topic_for_call(call):
        return

    chat_id = call.message.chat.id
    data = call.data

    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass

    if data == "ui_groups":
        send_locked(chat_id, GROUPS_TEXT, disable_web_page_preview=True, reply_markup=kb_useful_inline(),
                    fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "ui_protocol":
        try:
            bot.edit_message_text(
                "📝 <b>Протокол собрания</b>\nВыбери раздел 👇",
                chat_id,
                call.message.message_id,
                reply_markup=kb_protocol_inline()
            )
        except Exception:
            send_locked(chat_id, "📝 <b>Протокол собрания</b>\nВыбери раздел 👇", reply_markup=kb_protocol_inline(),
                        fallback_thread_id=get_thread_id_from_call(call))
        return


# ================== REMINDERS MENU CALLBACKS ==================
@bot.callback_query_handler(func=lambda call: call.data in {"rem_add", "rem_list"})
def reminders_menu_callbacks(call):
    if not in_allowed_topic_for_call(call):
        return

    chat_id = call.message.chat.id
    user_id = call.from_user.id
    data = call.data

    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass

    if data == "rem_add":
        clear_user_state(user_id)
        states[user_id] = {
            "step": "title",
            "chat_id": chat_id,
            "thread_id": get_thread_id_from_call(call)
        }
        send_locked(chat_id, "Ок! Введи <b>название</b> напоминания:", reply_markup=kb_cancel_inline(),
                    fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "rem_list":
        items = get_chat_reminders(chat_id)
        if not items:
            send_locked(chat_id, "Пока нет напоминаний в этом чате.", reply_markup=kb_reminders_inline(call.from_user),
                        fallback_thread_id=get_thread_id_from_call(call))
            return

        lines = ["📋 <b>Напоминания в этом чате</b>:"]
        for i, r in enumerate(items, 1):
            lines.append(f"{i}. <b>{r['title']}</b> — {format_event_dt(r['event_dt'])}")
        lines.append(f"\n🧹 Автоудаление: через {AUTO_DELETE_AFTER_HOURS} ч после события.")
        send_locked(chat_id, "\n".join(lines), reply_markup=kb_reminders_inline(call.from_user),
                    fallback_thread_id=get_thread_id_from_call(call))
        return


# ================== CALLBACKS (дата/время/отмена) ==================
@bot.callback_query_handler(
    func=lambda call: (
        call.data in {"cancel", "date_manual", "time_manual"} or
        call.data.startswith("date|") or
        call.data.startswith("time|")
    )
)
def callbacks_reminders(call):
    if not in_allowed_topic_for_call(call):
        return

    user_id = call.from_user.id
    chat_id = call.message.chat.id
    st = states.get(user_id)
    data = call.data

    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass

    if data == "cancel":
        clear_user_state(user_id)
        send_locked(chat_id, "Ок, отменил. Возвращаюсь в меню:", reply_markup=kb_reminders_inline(call.from_user),
                    fallback_thread_id=get_thread_id_from_call(call))
        return

    if not st or int(st.get("chat_id")) != int(chat_id):
        return

    if data.startswith("date|"):
        date_iso = data.split("|", 1)[1]
        st["date"] = date_iso
        st["step"] = "time_pick"
        try:
            bot.edit_message_text(
                "Дата выбрана ✅\nТеперь выбери <b>время</b>:",
                chat_id,
                call.message.message_id,
                reply_markup=build_time_picker()
            )
        except Exception:
            send_locked(chat_id, "Дата выбрана ✅\nТеперь выбери <b>время</b>:", reply_markup=build_time_picker(),
                        fallback_thread_id=get_thread_id_from_call(call))
        return

    if data == "date_manual":
        st["step"] = "date_manual"
        try:
            bot.edit_message_text(
                "Введи дату вручную: <b>31.12.2026</b> или <b>2026-12-31</b>",
                chat_id,
                call.message.message_id
            )
        except Exception:
            send_locked(chat_id, "Введи дату вручную: <b>31.12.2026</b> или <b>2026-12-31</b>",
                        fallback_thread_id=get_thread_id_from_call(call))
        return

    if data.startswith("time|"):
        time_hhmm = data.split("|", 1)[1]
        try:
            bot.edit_message_reply_markup(chat_id, call.message.message_id, reply_markup=None)
        except Exception:
            pass
        finalize_reminder(user_id, chat_id, time_hhmm)
        return

    if data == "time_manual":
        st["step"] = "time_manual"
        try:
            bot.edit_message_text(
                "Введи время вручную в формате <b>HH:MM</b> (например, <b>18:30</b>):",
                chat_id,
                call.message.message_id
            )
        except Exception:
            send_locked(chat_id, "Введи время вручную в формате <b>HH:MM</b> (например, <b>18:30</b>):",
                        fallback_thread_id=get_thread_id_from_call(call))
        return


# ================== CALLBACKS (сроки хранения) ==================
@bot.callback_query_handler(func=lambda call: call.data.startswith("storage_"))
def callbacks_storage(call):
    if not in_allowed_topic_for_call(call):
        return

    chat_id = call.message.chat.id
    user_id = call.from_user.id
    data = call.data

    try:
        bot.answer_callback_query(call.id)
    except Exception:
        pass

    if data == "storage_newsearch":
        states[user_id] = {"mode": "storage_search", "chat_id": chat_id, "thread_id": get_thread_id_from_call(call)}
        send_locked(chat_id, "🔎 Введи название продукта для поиска:", reply_markup=kb_storage_start(),
                    fallback_thread_id=get_thread_id_from_call(call))
        return

    if data.startswith("storage_pick|"):
        st = states.get(user_id, {})
        results = st.get("storage_results", [])
        try:
            idx = int(data.split("|", 1)[1])
        except Exception:
            idx = -1

        if not results or idx < 0 or idx >= len(results):
            send_locked(chat_id, "Не нашёл выбранный результат. Сделай новый поиск.", reply_markup=kb_storage_after_result(),
                        fallback_thread_id=get_thread_id_from_call(call))
            return

        row = results[idx]
        send_locked(chat_id, format_storage_row(row), reply_markup=kb_storage_after_result(),
                    fallback_thread_id=get_thread_id_from_call(call))
        clear_storage_mode(user_id)
        return


# ================== ТЕКСТОВЫЙ РОУТЕР (ТОЛЬКО КОГДА ЕСТЬ STATE) ==================
@bot.message_handler(func=lambda m: states.get(m.from_user.id) is not None, content_types=["text"])
def text_router(message):
    if not in_allowed_topic_for_message(message):
        return

    user_id = message.from_user.id
    st = states.get(user_id)
    if not st:
        return

    chat_id = st.get("chat_id")
    if int(chat_id) != int(message.chat.id):
        return

    # ====== режим поиска сроков хранения ======
    if st.get("mode") == "storage_search":
        query = (message.text or "").strip()
        if not query:
            send_locked(message.chat.id, "Введи название продукта текстом.", reply_markup=kb_storage_start(),
                        fallback_thread_id=get_thread_id_from_message(message))
            return

        if not STORAGE_READY:
            send_locked(message.chat.id, "База не загружена или пустая.", reply_markup=kb_storage_start(),
                        fallback_thread_id=get_thread_id_from_message(message))
            return

        results = storage_search(query, limit=12)
        if not results:
            send_locked(
                message.chat.id,
                f"Ничего не нашёл по запросу: <b>{query}</b>\n"
                "Попробуй другое слово или более короткий запрос.",
                reply_markup=kb_storage_start(),
                fallback_thread_id=get_thread_id_from_message(message)
            )
            return

        st["storage_results"] = results

        if len(results) == 1:
            send_locked(message.chat.id, format_storage_row(results[0]), reply_markup=kb_storage_after_result(),
                        fallback_thread_id=get_thread_id_from_message(message))
            clear_storage_mode(user_id)
            return

        send_locked(message.chat.id, f"Нашёл вариантов: <b>{len(results)}</b>\nВыбери нужный:", reply_markup=kb_storage_pick_list(results),
                    fallback_thread_id=get_thread_id_from_message(message))
        return

    # ====== сценарий напоминаний ======
    step = st.get("step")

    if step == "title":
        title = (message.text or "").strip()
        if not title:
            send_locked(message.chat.id, "Название не может быть пустым. Введи ещё раз:", reply_markup=kb_cancel_inline(),
                        fallback_thread_id=get_thread_id_from_message(message))
            return

        st["title"] = title
        st["step"] = "date_pick"
        send_locked(message.chat.id, "Выбери <b>дату</b>:", reply_markup=build_date_picker(),
                    fallback_thread_id=get_thread_id_from_message(message))
        return

    if step == "date_manual":
        raw = (message.text or "").strip()
        date_iso = None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(raw, fmt).date()
                date_iso = d.isoformat()
                break
            except ValueError:
                pass

        if not date_iso:
            send_locked(message.chat.id, "Не понял дату. Пример: <b>31.12.2026</b> или <b>2026-12-31</b>",
                        fallback_thread_id=get_thread_id_from_message(message))
            return

        st["date"] = date_iso
        st["step"] = "time_pick"
        send_locked(message.chat.id, "Теперь выбери <b>время</b>:", reply_markup=build_time_picker(),
                    fallback_thread_id=get_thread_id_from_message(message))
        return

    if step == "time_manual":
        raw = (message.text or "").strip()
        if not validate_time_hhmm(raw):
            send_locked(message.chat.id, "Не понял время. Пример: <b>18:30</b> (формат HH:MM)",
                        fallback_thread_id=get_thread_id_from_message(message))
            return

        finalize_reminder(user_id, message.chat.id, raw)
        return


def finalize_reminder(user_id: int, chat_id: int, time_hhmm: str) -> None:
    st = states.get(user_id)
    if not st:
        return

    title = st["title"]
    date_iso = st["date"]

    event_dt_naive = datetime.strptime(f"{date_iso} {time_hhmm}", "%Y-%m-%d %H:%M")
    event_dt = TZ.localize(event_dt_naive)

    if event_dt <= now_tz():
        send_locked(chat_id, "Это время уже в прошлом. Давай выберем заново дату/время.",
                    fallback_thread_id=st.get("thread_id"))
        st["step"] = "date_pick"
        send_locked(chat_id, "Выбери <b>дату</b>:", reply_markup=build_date_picker(),
                    fallback_thread_id=st.get("thread_id"))
        return

    # Для групп с закреплённой темой — всегда пишем туда.
    # Для групп без закрепления — используем тему, где создавали (если есть).
    thread_id = None
    allowed = get_allowed_thread_id(chat_id)
    if allowed is not None:
        thread_id = allowed
    else:
        thread_id = st.get("thread_id")

    rem = {
        "id": uuid.uuid4().hex,
        "chat_id": int(chat_id),
        "creator_id": int(user_id),
        "title": title,
        "event_dt": dt_to_iso(event_dt),
        "created_at": dt_to_iso(now_tz()),
        "thread_id": int(thread_id) if thread_id is not None else None
    }

    add_reminder_to_store(rem)
    schedule_reminder_jobs(rem)

    send_locked(
        chat_id,
        "✅ Напоминание добавлено!\n"
        f"<b>{title}</b>\n"
        f"📅 {event_dt.strftime('%d.%m.%Y %H:%M')}\n"
        "Я напомню <b>за 24 часа</b> и <b>за 1 час</b> до события.\n"
        f"🧹 Автоудаление: через <b>{AUTO_DELETE_AFTER_HOURS} ч</b> после события.\n\n"
        "Дальше что делаем?",
        reply_markup=kb_reminders_inline(),
        fallback_thread_id=thread_id
    )

    clear_user_state(user_id)


# ======= загрузка базы при старте =======
_count, _sheets = load_storage_db()


if __name__ == "__main__":
    print(f"🤖 Bot is running. TZ={TZ_NAME} | VERSION={BOT_VERSION}")
    print(f"🧊 Storage ready: {STORAGE_READY} | file: {STORAGE_SOURCE_PATH} | rows: {len(STORAGE_DB)}")
    bot.infinity_polling(skip_pending=True)
